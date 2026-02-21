"""Regenerate sample_data.xlsx with the new SOCS exam session schema."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Exam Sessions"

headers = [
    "Exam Title",
    "Room Number",
    "Date of Exam",
    "Time",
    "Program / Batch",
    "Semester",
    "Course Name",
    "Course Code",
    "Name of Evaluator",
    "No. of Students",
]

# Style header row
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="1E4BA0")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

# Sample data rows matching the SOCS label template
rows = [
    ["SoCS Mid Sem Exam March 2025", "2001", "05/03/2025", "02:00 PM - 04:00 PM", "BT-CSE-II-B1",  "2", "Computer Organization and Architecture", "CSEG1032", "Rajib Banerjee",  "29"],
    ["SoCS Mid Sem Exam March 2025", "2001", "05/03/2025", "02:00 PM - 04:00 PM", "BT-CSE-II-B10", "2", "Computer Organization and Architecture", "CSEG1032", "Avishek Majumder", "4"],
    ["SoCS Mid Sem Exam March 2025", "2002", "05/03/2025", "02:00 PM - 04:00 PM", "BT-CSE-III-A1", "4", "Data Structures",                        "CSEG2011", "Dr. Priya Sharma",  "35"],
    ["SoCS Mid Sem Exam March 2025", "2003", "06/03/2025", "09:00 AM - 11:00 AM", "BT-CSE-IV-B1",  "6", "Machine Learning",                       "CSEG4021", "Dr. Anil Verma",   "30"],
    ["SoCS Mid Sem Exam March 2025", "2003", "06/03/2025", "09:00 AM - 11:00 AM", "BT-CSE-IV-B2",  "6", "Machine Learning",                       "CSEG4021", "Dr. Anil Verma",   "28"],
]

for row in rows:
    ws.append(row)

# Auto-size columns
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

out = os.path.join(os.path.dirname(__file__), 'Samples', 'sample_data.xlsx')
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: {len(rows)}, Columns: {len(headers)}")
